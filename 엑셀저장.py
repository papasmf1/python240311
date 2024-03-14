import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import random

# 엑셀 파일 생성
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Products'

# 열 제목 설정
columns = ['제품ID', '제품이름', '제품가격']
for col_num, column_title in enumerate(columns, start=1):
    sheet[get_column_letter(col_num) + '1'] = column_title
    sheet[get_column_letter(col_num) + '1'].font = Font(bold=True)

# 전자제품 데이터 생성 및 입력
for row_num in range(2, 102):  # 100개의 제품 데이터 생성
    product_id = row_num - 1
    product_name = f'제품{product_id}'
    product_price = random.randint(10000, 1000000)
    
    sheet['A' + str(row_num)] = product_id
    sheet['B' + str(row_num)] = product_name
    sheet['C' + str(row_num)] = product_price

# 엑셀 파일 저장
wb.save('products.xlsx')
print("전자제품 데이터가 생성되어 'products.xlsx' 파일에 저장되었습니다.")
